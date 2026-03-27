"""使用 openpyxl 读取 xlsx 并解析为 TableSchema 的模块。"""

from __future__ import annotations

import logging
import os

import openpyxl

from table_parser.helper import format_class_name, is_legal_file_name
from table_parser.types import BelongType, FieldInfo, FieldType, TableSchema

logger = logging.getLogger(__name__)

COMMENT_ROW = 1
BELONG_ROW = 2
TYPE_ROW = 3
FIELD_NAME_ROW = 4
DATA_START_ROW = 5

# read_only 模式下需用 iter_rows 扫描列；一般配置表远小于此上限。
_MAX_HEADER_COL = 1024


class ExcelParser:
    """将符合约定表头的 xlsx 首个工作表解析为 TableSchema。"""

    def parse(self, file_path: str) -> TableSchema:
        """解析 xlsx 文件，返回 TableSchema。

        :param file_path: xlsx 路径
        :raises ValueError: 文件名不合法、类型/归属无法识别、无字段或校验失败
        """
        base = os.path.basename(file_path)
        file_name, ext = os.path.splitext(base)
        if ext.lower() != ".xlsx":
            logger.warning("文件扩展名不是 .xlsx：%s", file_path)

        if not is_legal_file_name(file_name):
            raise ValueError(
                f"表文件名不合法（须为小写字母与下划线，且首尾不能为下划线）：{file_name!r}"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = wb.worksheets[0]
            fields = self._read_header(sheet)
            key_field = self._find_key_field(fields)
            rows = self._read_data_rows(sheet, fields)
            class_name = format_class_name(file_name)
            schema = TableSchema(
                file_name=file_name,
                class_name=class_name,
                fields=fields,
                key_field=key_field,
                rows=rows,
            )
            self._validate(schema)
            return schema
        finally:
            wb.close()

    def _read_header(self, sheet) -> list[FieldInfo]:
        """读取前四行表头，按列生成 FieldInfo 列表（遇空字段名列结束）。"""
        rows_list = list(
            sheet.iter_rows(
                min_row=COMMENT_ROW,
                max_row=FIELD_NAME_ROW,
                min_col=1,
                max_col=_MAX_HEADER_COL,
                values_only=True,
            )
        )
        if len(rows_list) < FIELD_NAME_ROW:
            # 不足四行时按缺失补 None
            while len(rows_list) < FIELD_NAME_ROW:
                rows_list.append(())

        def cell(row_index: int, col: int):
            r = rows_list[row_index - 1]
            if col - 1 < len(r):
                return r[col - 1]
            return None

        fields: list[FieldInfo] = []
        col = 1
        while col <= _MAX_HEADER_COL:
            raw_name = cell(FIELD_NAME_ROW, col)
            if raw_name is None or str(raw_name).strip() == "":
                break

            name = str(raw_name).strip()
            comment = cell(COMMENT_ROW, col)
            comment_s = "" if comment is None else str(comment).strip()

            belong = self._parse_belong(cell(BELONG_ROW, col))
            field_type = self._parse_field_type(cell(TYPE_ROW, col))

            fields.append(
                FieldInfo(
                    name=name,
                    field_type=field_type,
                    belong=belong,
                    comment=comment_s,
                    column_index=col,
                )
            )
            col += 1

        return fields

    def _parse_belong(self, raw) -> BelongType:
        """解析归属单元格：空为 NONE，否则大写后匹配 BelongType。"""
        if raw is None:
            return BelongType.NONE
        s = str(raw).strip()
        if not s:
            return BelongType.NONE
        u = s.upper()
        for bt in BelongType:
            if bt.value == u:
                return bt
        raise ValueError(f"无法识别的归属（Belong）：{raw!r}，应为 A/C/S/K/N 之一")

    def _parse_field_type(self, raw) -> FieldType:
        """解析类型单元格：小写后匹配 FieldType。"""
        if raw is None or str(raw).strip() == "":
            raise ValueError("已定义字段的类型不能为空")
        s = str(raw).strip().lower()
        for ft in FieldType:
            if ft.value == s:
                return ft
        raise ValueError(
            f"无法识别的类型：{raw!r}，支持："
            + ", ".join(ft.value for ft in FieldType)
        )

    def _find_key_field(self, fields: list[FieldInfo]) -> FieldInfo:
        """确定主键列：优先 KEY 归属，其次名为 Id 的列，否则第一列。"""
        if not fields:
            raise ValueError("表没有任何有效字段列")

        key_candidates = [f for f in fields if f.belong == BelongType.KEY]
        if len(key_candidates) > 1:
            logger.warning(
                "存在多个 KEY 归属列，将使用第一个作为主键：%s",
                ", ".join(f.name for f in key_candidates),
            )
        if key_candidates:
            return key_candidates[0]

        for f in fields:
            if f.name == "Id":
                return f

        return fields[0]

    def _read_data_rows(self, sheet, fields: list[FieldInfo]) -> list[list[str]]:
        """从第 DATA_START_ROW 行起读取数据；空行跳过；单元格 None 转为空串。"""
        if not fields:
            return []

        min_c = min(f.column_index for f in fields)
        max_c = max(f.column_index for f in fields)

        out: list[list[str]] = []
        for row in sheet.iter_rows(
            min_row=DATA_START_ROW,
            min_col=min_c,
            max_col=max_c,
            values_only=True,
        ):
            vals: list[str] = []
            for f in fields:
                idx = f.column_index - min_c
                if idx < 0 or idx >= len(row):
                    cell_v = None
                else:
                    cell_v = row[idx]
                vals.append("" if cell_v is None else str(cell_v))

            if all(v == "" for v in vals):
                continue
            out.append(vals)

        return out

    def _validate(self, schema: TableSchema) -> None:
        """校验字段名重复（非 NONE 列、忽略大小写）及主键存在性。"""
        if schema.key_field is None:
            raise ValueError("主键字段 key_field 不能为空")

        active = [f for f in schema.fields if f.belong is not BelongType.NONE]
        seen: dict[str, str] = {}
        for f in active:
            lower = f.name.lower()
            if lower in seen:
                raise ValueError(
                    f"非 NONE 字段存在重名（忽略大小写）：{seen[lower]!r} 与 {f.name!r}"
                )
            seen[lower] = f.name

        if not schema.fields:
            logger.warning("表 %s 无字段定义", schema.file_name)
